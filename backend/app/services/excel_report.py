"""
Regenera reporte_clasificacion.xlsx a partir del estado en la base de datos.
Se puede llamar en cualquier momento — no depende de que ningún proceso en
memoria haya sobrevivido, solo lee lo que ya está persistido en SQLite.

Escritura atómica: escribe a un .tmp y hace os.replace() al final, así si el
proceso se corta a mitad de la escritura, el Excel anterior queda intacto.
"""

import json
import os
from pathlib import Path

from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from app.config import get_settings
from app.models.job import Job, JobStatus, SummaryStatus

settings = get_settings()

EXCEL_FILENAME = "reporte_clasificacion.xlsx"

_HEADER_FONT = Font(bold=True)


def _autosize(ws, widths: dict[int, int]):
    for col_idx, width in widths.items():
        ws.column_dimensions[chr(64 + col_idx)].width = width


def _write_header(ws, headers: list[str]):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT


def regenerate_excel(db: Session) -> str:
    jobs = db.query(Job).order_by(Job.filename).all()

    wb = Workbook()

    # Hoja 1: documentos MD generados (salida del OCR)
    ws1 = wb.active
    ws1.title = "Documentos MD"
    _write_header(ws1, ["Documento original", "Archivo .md", "Estado OCR", "Confianza promedio", "Completado"])
    for job in jobs:
        md_name = Path(job.output_path).name if job.output_path else ""
        completed = job.completed_at.strftime("%Y-%m-%d %H:%M") if job.completed_at else ""
        ws1.append([
            job.filename,
            md_name,
            job.status.value if job.status else "",
            round(job.avg_confidence or 0.0, 3),
            completed,
        ])
    _autosize(ws1, {1: 35, 2: 35, 3: 14, 4: 16, 5: 18})

    # Hoja 2: resúmenes MD generados
    ws2 = wb.create_sheet("Resúmenes MD")
    _write_header(ws2, ["Documento original", "Archivo resumen .md", "Estado resumen", "Error"])
    for job in jobs:
        summary_name = Path(job.summary_md_path).name if job.summary_md_path else ""
        ws2.append([
            job.filename,
            summary_name,
            job.summary_status.value if job.summary_status else "",
            (job.summary_error or "")[:300],
        ])
    _autosize(ws2, {1: 35, 2: 35, 3: 16, 4: 50})

    # Hoja 3: top-5 de clasificación (una fila por documento+categoría)
    ws3 = wb.create_sheet("Top 5 Clasificación")
    _write_header(ws3, ["Documento original", "Rank", "Categoría", "Score", "Justificación"])
    for job in jobs:
        if not job.classification_json:
            continue
        try:
            data = json.loads(job.classification_json)
        except (json.JSONDecodeError, TypeError):
            continue
        for rank, item in enumerate(data.get("clasificacion_top5", []), start=1):
            ws3.append([
                job.filename,
                rank,
                item.get("categoria", ""),
                item.get("score", ""),
                item.get("justificacion", ""),
            ])
    _autosize(ws3, {1: 35, 2: 6, 3: 25, 4: 8, 5: 60})

    output_dir = Path(settings.output_path)
    output_dir.mkdir(parents=True, exist_ok=True)
    final_path = output_dir / EXCEL_FILENAME
    tmp_path = output_dir / f".{EXCEL_FILENAME}.tmp"

    wb.save(str(tmp_path))
    os.replace(str(tmp_path), str(final_path))

    logger.info(f"Excel regenerado: {final_path} ({len(jobs)} documentos)")
    return str(final_path)
